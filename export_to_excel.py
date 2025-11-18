#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将分割好的图片导出到 Excel，方便打印
"""

import os
import sys
from pathlib import Path
import argparse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image


def export_to_excel(image_dir, output_file=None, num_columns=3, column_width=25, 
                    row_height=150, page_break_rows=None):
    """
    将分割好的列图片导入到 Excel，每页显示指定列数
    
    参数:
        image_dir: 包含分割图片的目录
        output_file: 输出的 Excel 文件名
        num_columns: 每页显示的列数
        column_width: Excel 列宽（单位：字符）
        row_height: Excel 行高（单位：磅）
        page_break_rows: 每多少行插入分页符（None=自动）
    """
    try:
        image_dir = Path(image_dir)
        
        # 查找所有列图片（格式：*_列1.png, *_列2.png 等）
        column_images = sorted([f for f in image_dir.glob("*_列*.png")])
        
        if not column_images:
            print(f"❌ 在目录 {image_dir} 中未找到列图片")
            print("提示：请先运行 split_long_image.py 生成列图片")
            return False
        
        # 按文件名分组（同一张图的不同列）
        from collections import defaultdict
        image_groups = defaultdict(list)
        
        for img_path in column_images:
            # 提取基础文件名（去掉_列N部分）
            base_name = img_path.stem.rsplit('_列', 1)[0]
            image_groups[base_name].append(img_path)
        
        print(f"找到 {len(image_groups)} 张图片，共 {len(column_images)} 列")
        
        # 创建 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "打印预览"
        
        # 设置打印选项
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # 横向
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = False
        
        # 设置页边距（单位：英寸）
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        
        # 设置列宽
        for col in range(1, num_columns + 1):
            ws.column_dimensions[get_column_letter(col)].width = column_width
        
        current_row = 1
        
        # 处理每组图片
        for idx, (base_name, images) in enumerate(sorted(image_groups.items())):
            print(f"\n处理: {base_name}")
            print(f"  包含 {len(images)} 列")
            
            # 获取第一张图片的高度来计算行高
            first_img = Image.open(images[0])
            img_width, img_height = first_img.size
            
            # 根据图片宽高比调整行高
            aspect_ratio = img_height / img_width
            calculated_row_height = row_height * aspect_ratio
            
            # 设置当前行的行高
            ws.row_dimensions[current_row].height = min(calculated_row_height, 800)
            
            # 插入图片到各列
            for col_idx, img_path in enumerate(sorted(images)[:num_columns], start=1):
                # 读取图片
                img = XLImage(str(img_path))
                
                # 调整图片大小以适应单元格
                # Excel 中的单位转换：列宽(字符) * 7 ≈ 像素
                cell_width_px = column_width * 7
                cell_height_px = calculated_row_height * 1.33  # 磅转像素
                
                # 保持宽高比缩放
                scale_w = cell_width_px / img.width
                scale_h = cell_height_px / img.height
                scale = min(scale_w, scale_h) * 0.95  # 0.95 留一点边距
                
                img.width = int(img.width * scale)
                img.height = int(img.height * scale)
                
                # 插入图片到单元格
                cell = ws.cell(row=current_row, column=col_idx)
                ws.add_image(img, cell.coordinate)
                
                print(f"  ✓ 已插入 {img_path.name} 到 {cell.coordinate}")
            
            current_row += 1
            
            # 插入分页符（每处理完一组图片后）
            if page_break_rows and (idx + 1) % page_break_rows == 0:
                ws.row_breaks.append(current_row - 1)
                print(f"  📄 已在第 {current_row - 1} 行后插入分页符")
        
        # 保存 Excel 文件
        if output_file is None:
            output_file = image_dir / "打印预览.xlsx"
        else:
            output_file = Path(output_file)
        
        wb.save(output_file)
        print(f"\n✅ 成功！Excel 文件已保存: {output_file.absolute()}")
        print(f"\n📋 打印说明:")
        print(f"  1. 打开 {output_file.name}")
        print(f"  2. 文件 → 打印（⌘ + P）")
        print(f"  3. 确认纸张方向为「横向」")
        print(f"  4. 选择「适合页面」或「缩放到纸张大小」")
        print(f"  5. 点击打印")
        
        return True
        
    except Exception as e:
        print(f"❌ 错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def main():
    parser = argparse.ArgumentParser(
        description='将分割好的图片导出到 Excel，方便打印',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # 将 output 目录中的列图片导出到 Excel（3列布局）
  python export_to_excel.py output/
  
  # 指定输出文件名
  python export_to_excel.py output/ -o 我的打印文件.xlsx
  
  # 自定义列数和尺寸
  python export_to_excel.py output/ -c 3 --column-width 30 --row-height 200
  
  # 每1组图片后插入分页符
  python export_to_excel.py output/ --page-break 1
        """
    )
    
    parser.add_argument('input_dir', help='包含分割列图片的目录（通常是 output/）')
    parser.add_argument('-o', '--output', default=None,
                        help='输出的 Excel 文件名（默认: 打印预览.xlsx）')
    parser.add_argument('-c', '--columns', type=int, default=3,
                        help='每页显示的列数（默认: 3）')
    parser.add_argument('--column-width', type=float, default=25,
                        help='Excel 列宽，单位：字符（默认: 25）')
    parser.add_argument('--row-height', type=float, default=150,
                        help='Excel 行高，单位：磅（默认: 150）')
    parser.add_argument('--page-break', type=int, default=None,
                        help='每多少组图片插入一个分页符（默认: 不插入）')
    
    args = parser.parse_args()
    
    # 检查输入目录
    input_path = Path(args.input_dir)
    if not input_path.exists():
        print(f"❌ 错误: 目录不存在: {args.input_dir}")
        sys.exit(1)
    
    if not input_path.is_dir():
        print(f"❌ 错误: {args.input_dir} 不是一个目录")
        sys.exit(1)
    
    # 执行导出
    success = export_to_excel(
        args.input_dir,
        args.output,
        args.columns,
        args.column_width,
        args.row_height,
        args.page_break
    )
    
    if not success:
        sys.exit(1)


if __name__ == "__main__":
    main()
